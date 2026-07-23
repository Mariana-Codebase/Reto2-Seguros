"""
perfilar_base.py
----------------
Procesa la base real de afiliados (`Usos_Productos_Afiliados_SIN_ID.xlsx`,
500.000 registros, hoja 'in') y genera los insumos precalculados que la demo
usa sin necesitar el archivo:

  data/mapa_segmentos.json     Correspondencia DOCUMENTADA de las etiquetas
                               anonimizadas (SIGMA, LAMBDA, RHO...) con su
                               significado, construida por coincidencia de
                               participaciones con la distribución pública del
                               insumo inicial del reto y VALIDADA con evidencia
                               interna (edad, salario, marcas de consumo).
                               Es editable: si Colsubsidio entrega el
                               diccionario oficial, se corrige aquí y listo.
  data/propension_stats.json   Agregados de toda la base + a cuántos afiliados
                               el motor recomienda cada producto como 1ª opción.
  data/afiliados_demo.json     Muestra curada de perfiles reales (solo SERIE +
                               variables) cubriendo los arquetipos más
                               distintos entre sí.

Uso:
  python scripts/perfilar_base.py "C:/ruta/Usos_Productos_Afiliados_SIN_ID.xlsx"

La base NO se versiona en el repo. Los JSON generados sí: son agregados,
mapeos documentados o registros anonimizados (la base no trae nombres).
"""

from __future__ import annotations

import csv
import json
import pathlib
import sys
from collections import Counter, defaultdict
from typing import Any, Iterator

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from app import propension  # noqa: E402

BASE = pathlib.Path(__file__).resolve().parents[1]
OUT_MAPA = BASE / "data" / "mapa_segmentos.json"
OUT_STATS = BASE / "data" / "propension_stats.json"
OUT_DEMO = BASE / "data" / "afiliados_demo.json"

MARCAS = ["HOTELES", "PISCILAGO", "DROGUERIA", "AGENCIAS", "VIVIENDA"]

# ---------------------------------------------------------------------------
# Distribución de referencia: participaciones publicadas en el insumo inicial
# del reto (mismo universo de afiliados, etiquetas en claro). El mapeo asigna
# a cada etiqueta anonimizada el significado cuya participación coincide.
# ---------------------------------------------------------------------------
REFERENCIA = {
    "CATEGORIA": [
        ("A", "Categoría A del subsidio familiar (ingresos hasta 2 SMLV)", 75.8),
        ("B", "Categoría B del subsidio familiar (ingresos 2 a 4 SMLV)", 14.5),
        ("C", "Categoría C del subsidio familiar (ingresos > 4 SMLV)", 8.8),
        ("sin_dato", "Sin categoría registrada", 0.9),
    ],
    "SEGMENTO_GRUPO_FAMILIAR": [
        ("sin_grupo", "Afiliado sin grupo familiar registrado", 58.0),
        ("monoparental", "Familia monoparental", 23.5),
        ("nuclear", "Familia nuclear integral (con hijos)", 9.4),
        ("pareja", "Pareja conyugal sin hijos registrados", 5.5),
        ("monoparental_ampliada", "Familia monoparental ampliada", 2.2),
        ("sin_dato", "Sin segmento registrado", 1.5),
        ("nuclear_ampliada", "Familia nuclear ampliada", 0.01),
    ],
    "SEGMENTO_POBLACIONAL": [
        ("basico", "Segmento Básico (ingresos/edad/PAC nivel básico)", 48.8),
        ("medio", "Segmento Medio", 26.8),
        ("joven", "Segmento Joven", 23.1),
        ("sin_dato", "Sin segmento registrado", 0.9),
        ("alto", "Segmento Alto", 0.4),
    ],
    "EMPRESA_FOCO": [
        ("no_foco", "Sin empresa foco asociada", 83.2),
        ("foco", "Afiliado de empresa foco Colsubsidio", 16.8),
    ],
    "PIRAMIDE_NUEVA": [
        ("micro", "Empresa micro transaccional", 32.5),
        ("medianas", "Empresas medianas", 20.1),
        ("grandes", "Empresas grandes", 20.0),
        ("empresarial_top", "Empresarial top", 11.6),
        ("facultativo", "Afiliado facultativo (6.1)", 5.2),
        ("empresarial_estandar", "Empresarial estándar", 4.0),
        ("pensionado", "Pensionado (6.3)", 2.8),
        ("independiente", "Trabajador independiente (6.2)", 1.6),
        ("sin_dato", "Sin clasificación registrada", 1.3),
        ("micro_colsubsidio", "Micro transaccional Colsubsidio", 0.9),
    ],
}


def leer_filas(path: str) -> Iterator[dict[str, Any]]:
    """Itera la base como dicts, soportando .xlsx (hoja 'in') y .csv (';')."""
    p = pathlib.Path(path)
    if p.suffix.lower() == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(rows)]
        for r in rows:
            yield dict(zip(header, r))
        wb.close()
    else:
        with open(p, encoding="utf-8-sig", newline="") as f:
            yield from csv.DictReader(f, delimiter=";")


# ---------------------------------------------------------------------------
# Fase 1: distribuciones + evidencia para el mapa
# ---------------------------------------------------------------------------
def fase1(path: str) -> dict[str, Any]:
    n = 0
    dist = {d: Counter() for d in REFERENCIA}
    marcas = Counter()
    edad = Counter()
    salario = Counter()
    empresas = Counter()
    # evidencia: por etiqueta, distribución de edad / salario / marcas
    ev_edad = {d: defaultdict(Counter) for d in ("PIRAMIDE_NUEVA", "SEGMENTO_POBLACIONAL")}
    ev_sal = {d: defaultdict(Counter) for d in ("CATEGORIA", "SEGMENTO_POBLACIONAL")}
    ev_marca = {d: defaultdict(Counter) for d in REFERENCIA}

    for row in leer_filas(path):
        a = propension.parse_afiliado(row)
        n += 1
        edad[a["rango_edad"] or "(sin dato)"] += 1
        salario[a["rango_salarial"] or "(sin dato)"] += 1
        empresas[a["empresa"] or "(sin dato)"] += 1
        for m, v in a["marcas"].items():
            if v:
                marcas[m] += 1
        for dim, campo in [("CATEGORIA", "categoria"),
                           ("SEGMENTO_GRUPO_FAMILIAR", "segmento_familiar"),
                           ("SEGMENTO_POBLACIONAL", "segmento_poblacional"),
                           ("PIRAMIDE_NUEVA", "piramide"),
                           ("EMPRESA_FOCO", "empresa")]:
            lab = a[campo] or "(vacio)"
            dist[dim][lab] += 1
            ev = ev_marca[dim][lab]
            ev["_n"] += 1
            for m, v in a["marcas"].items():
                if v:
                    ev[m] += 1
            if dim in ev_edad:
                ev_edad[dim][lab][a["rango_edad"] or "?"] += 1
            if dim in ev_sal:
                ev_sal[dim][lab][a["rango_salarial"] or "?"] += 1
        if n % 100_000 == 0:
            print(f"  fase 1 ... {n:,} filas")

    return {"n": n, "dist": dist, "marcas": marcas, "edad": edad, "salario": salario,
            "empresas": empresas, "ev_edad": ev_edad, "ev_sal": ev_sal, "ev_marca": ev_marca}


def construir_mapa(st: dict[str, Any]) -> dict[str, Any]:
    """Asigna significado a cada etiqueta por coincidencia de participaciones
    con la referencia pública, y anexa la evidencia interna que lo respalda."""
    n = st["n"]
    mapa: dict[str, Any] = {
        "_metodo": (
            "Correspondencia por coincidencia de participaciones con la distribución "
            "pública del insumo inicial del reto (mismo universo de afiliados), "
            "validada con evidencia interna de la propia base (edad, salario y marcas "
            "de consumo por etiqueta). Editable: ante el diccionario oficial de "
            "Colsubsidio, corregir aquí; el motor no cambia."
        ),
    }
    for dim, ref in REFERENCIA.items():
        etiquetas = sorted(st["dist"][dim].items(), key=lambda kv: -kv[1])
        entradas = {}
        for (lab, cnt), (clave, significado, ref_share) in zip(etiquetas, ref):
            share = 100 * cnt / n
            diff = abs(share - ref_share)
            confianza = "alta" if diff <= 2.5 else ("media" if diff <= 6 else "baja")
            evidencia = {
                "participacion_base": round(share, 1),
                "participacion_referencia": ref_share,
            }
            # Validaciones internas específicas
            if dim == "PIRAMIDE_NUEVA" and clave == "pensionado":
                ed = st["ev_edad"][dim][lab]
                tot = sum(ed.values()) or 1
                evidencia["pct_mayor_55"] = round(100 * ed.get("Mayor de 55 años", 0) / tot, 1)
            if dim == "SEGMENTO_POBLACIONAL" and clave == "joven":
                ed = st["ev_edad"][dim][lab]
                tot = sum(ed.values()) or 1
                evidencia["pct_20_35"] = round(100 * ed.get("20 a 35 años", 0) / tot, 1)
            if dim == "CATEGORIA" and clave in ("A", "B", "C"):
                sal = st["ev_sal"][dim][lab]
                tot = sum(sal.values()) or 1
                bajos = sum(v for k, v in sal.items()
                            if k in ("Menor al SMLV", "Entre 1 y 1.5 SMLV", "Entre 1.5 y 2 SMLV"))
                evidencia["pct_hasta_2_smlv"] = round(100 * bajos / tot, 1)
            entradas[lab] = {"clave": clave, "significado": significado,
                             "confianza": confianza, "evidencia": evidencia}
        # etiquetas sobrantes sin pareja en la referencia
        for lab, cnt in etiquetas[len(ref):]:
            entradas[lab] = {"clave": "", "significado": "(sin correspondencia)",
                             "confianza": "baja",
                             "evidencia": {"participacion_base": round(100 * cnt / n, 1)}}
        mapa[dim] = {"etiquetas": entradas}
    return mapa


# ---------------------------------------------------------------------------
# Fase 2: motor sobre toda la base + muestra demo
# ---------------------------------------------------------------------------
ARQUETIPOS = [
    ("salud_activa", "Cabeza de hogar monoparental que compra en droguerías",
     lambda a, s: a["marcas"]["drogueria"] and s("SEGMENTO_GRUPO_FAMILIAR", a["segmento_familiar"]) == "monoparental", 2),
    ("viajera", "Compra en agencias de viajes Colsubsidio",
     lambda a, s: a["marcas"]["agencias"], 2),
    ("hotelera", "Se hospeda en hoteles Colsubsidio",
     lambda a, s: a["marcas"]["hoteles"] and not a["marcas"]["agencias"], 2),
    ("hogar_nuevo", "Usó el servicio de vivienda Colsubsidio",
     lambda a, s: a["marcas"]["vivienda"], 2),
    ("pensionado", "Pensionado mayor de 55 años",
     lambda a, s: s("PIRAMIDE_NUEVA", a["piramide"]) == "pensionado"
     and a["rango_edad"] == "Mayor de 55 años", 2),
    ("joven_solo", "Joven sin grupo familiar, ingresos hasta 1.5 SMLV",
     lambda a, s: s("SEGMENTO_GRUPO_FAMILIAR", a["segmento_familiar"]) == "sin_grupo"
     and a["rango_edad"] == "20 a 35 años"
     and a["rango_salarial"] in ("Menor al SMLV", "Entre 1 y 1.5 SMLV"), 2),
    ("independiente", "Trabajador independiente",
     lambda a, s: s("PIRAMIDE_NUEVA", a["piramide"]) == "independiente", 2),
    ("familia_alta", "Familia nuclear con ingresos superiores a 6 SMLV",
     lambda a, s: s("SEGMENTO_GRUPO_FAMILIAR", a["segmento_familiar"]) == "nuclear"
     and a["rango_salarial"] in ("Entre 6 y 8 SMLV", "Entre 8 y 10 SMLV",
                                  "Entre 10 y 20 SMLV", "Entre 20 y 30 SMLV"), 2),
]


def fase2(path: str) -> tuple[Counter, dict[str, Counter], dict[str, list[dict]]]:
    top1 = Counter()
    reglas_top: dict[str, Counter] = {}
    demo: dict[str, list[dict]] = {k: [] for k, *_ in ARQUETIPOS}
    n = 0
    for row in leer_filas(path):
        a = propension.parse_afiliado(row)
        n += 1
        p = propension.perfilar(a, top=1)
        if p["productos"]:
            prod = p["productos"][0]
            top1[prod["producto_id"]] += 1
            rz = reglas_top.setdefault(prod["producto_id"], Counter())
            for r in prod["razones"]:
                rz[r["regla"]] += 1
        for clave, _desc, filtro, cupo in ARQUETIPOS:
            if len(demo[clave]) < cupo:
                try:
                    ok = filtro(a, propension._sentido)
                except Exception:  # noqa: BLE001
                    ok = False
                if ok:
                    demo[clave].append(a)
                    break
        if n % 100_000 == 0:
            print(f"  fase 2 ... {n:,} filas")
    return top1, reglas_top, demo


def main(path: str) -> None:
    print("Fase 1: distribuciones y evidencia del mapa")
    st = fase1(path)
    n = st["n"]
    mapa = construir_mapa(st)
    OUT_MAPA.write_text(json.dumps(mapa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {OUT_MAPA}")
    # recargar el mapa en el motor antes de la fase 2
    propension._mapa_cache = None

    print("Fase 2: motor de propensión sobre toda la base + muestra demo")
    top1, reglas_top, demo = fase2(path)

    stats = {
        "total_afiliados": n,
        "fuente": "Usos_Productos_Afiliados_SIN_ID.xlsx (base anonimizada Colsubsidio, sin nombres ni cédulas)",
        "recomendacion_top1": dict(top1.most_common()),
        "reglas_dominantes_por_producto": {p: dict(c.most_common(5)) for p, c in reglas_top.items()},
        "distribuciones": {
            "rango_edad": dict(st["edad"].most_common()),
            "rango_salarial": dict(st["salario"].most_common()),
            "marcas_consumo": dict(st["marcas"].most_common()),
            **{d: dict(st["dist"][d].most_common()) for d in REFERENCIA},
        },
        "empresas_top": dict(st["empresas"].most_common(5)),
    }
    OUT_STATS.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {OUT_STATS} ({n:,} afiliados)")

    perfiles = []
    for clave, desc, _f, _c in ARQUETIPOS:
        for a in demo[clave]:
            perfiles.append({**a, "arquetipo": clave, "arquetipo_desc": desc,
                             "propension": propension.perfilar(a)})
    OUT_DEMO.write_text(json.dumps(perfiles, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {OUT_DEMO} ({len(perfiles)} perfiles demo, solo SERIE + variables)")

    print("\nTop-1 del motor sobre la base:")
    for pid, c in top1.most_common():
        print(f"  {pid}: {c:,} ({100*c/n:.1f}%)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python scripts/perfilar_base.py <ruta a la base (.xlsx o .csv)>")
    main(sys.argv[1])
